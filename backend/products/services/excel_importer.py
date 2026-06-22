"""Import products, brands, and source URLs from the Excel file.

Excel structure (column positions are detected, not hard-coded):
    - A header row contains the labels: MODEL, BRAND, and one column per
      website (e.g. 'digikala.com', 'techno life', 'mobile 140').
    - Each following row has the model name under MODEL, the brand under
      BRAND, and the product URL under each website column.

The header row is located by finding the row that contains a cell equal to
"MODEL" (case-insensitive). Everything is keyed off the labels in that row,
so inserting a new column (like BRAND) or shifting rows doesn't break import.

The importer is idempotent: re-importing updates URLs in place and does not
create duplicate ProductSource rows.

SYNC MODE (sync=True): after importing, any Product whose model_name is NOT in
the spreadsheet is DELETED, so the database becomes an exact mirror of the
Excel. Use this when the Excel is the single source of truth. Without it, the
importer only adds/updates and never removes (the safe default).
"""
from __future__ import annotations

import logging
from dataclasses import dataclass, field
from pathlib import Path
from typing import IO

from django.db import transaction
from openpyxl import load_workbook

from ..models import Brand, Product, ProductSource

logger = logging.getLogger(__name__)

# Labels we recognize in the header row (compared case-insensitively).
MODEL_LABEL = "model"
BRAND_LABEL = "brand"
# Any header cell that isn't MODEL or BRAND (and is non-empty) is treated as
# a website column.


@dataclass
class ImportReport:
    products_created: int = 0
    products_updated: int = 0
    products_deleted: int = 0
    sources_created: int = 0
    sources_updated: int = 0
    sources_deleted: int = 0
    brands_created: int = 0
    skipped_rows: int = 0
    website_columns: list[str] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)

    def as_dict(self) -> dict:
        return {
            "products_created": self.products_created,
            "products_updated": self.products_updated,
            "products_deleted": self.products_deleted,
            "sources_created": self.sources_created,
            "sources_updated": self.sources_updated,
            "sources_deleted": self.sources_deleted,
            "brands_created": self.brands_created,
            "skipped_rows": self.skipped_rows,
            "website_columns": self.website_columns,
            "warnings": self.warnings,
        }


def _find_header_and_columns(
    rows: list[list],
) -> tuple[int, int, int | None, dict[int, str]] | None:
    """Locate the header row and resolve column indices.

    Returns (header_index, model_col, brand_col_or_None, website_columns)
    where website_columns maps column-index -> website name. Returns None if
    no row containing a "MODEL" label is found.
    """
    for index, row in enumerate(rows):
        model_col = None
        brand_col = None
        website_columns: dict[int, str] = {}
        for col_index, cell in enumerate(row):
            if not isinstance(cell, str):
                continue
            label = cell.strip().lower()
            if not label:
                continue
            if label == MODEL_LABEL:
                model_col = col_index
            elif label == BRAND_LABEL:
                brand_col = col_index
            else:
                # Treat any other non-empty header label as a website column.
                website_columns[col_index] = cell.strip()
        # A valid header must have a MODEL column and at least one website.
        if model_col is not None and website_columns:
            return index, model_col, brand_col, website_columns
    return None


def _normalize_url(value) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if not text:
        return ""
    if not text.startswith(("http://", "https://")):
        text = "https://" + text.lstrip("/")
    return text


@transaction.atomic
def import_excel(file_obj: str | Path | IO[bytes], sync: bool = False) -> ImportReport:
    """Parse the spreadsheet and upsert Brand + Product + ProductSource rows.

    If sync=True, products not present in the spreadsheet are deleted so the
    database exactly mirrors the Excel.
    """
    report = ImportReport()

    workbook = load_workbook(file_obj, data_only=True, read_only=True)
    sheet = workbook.active

    rows = [list(row) for row in sheet.iter_rows(values_only=True)]
    if not rows:
        report.warnings.append("Spreadsheet is empty")
        return report

    resolved = _find_header_and_columns(rows)
    if resolved is None:
        report.warnings.append(
            "Could not locate a header row containing a 'MODEL' column"
        )
        return report

    header_index, model_col, brand_col, website_columns = resolved
    report.website_columns = list(website_columns.values())

    # Track every model name we see in the Excel (for sync-mode deletion).
    seen_model_names: set[str] = set()

    # Cache brands we touch this run so we only hit the DB once per name.
    brand_cache: dict[str, Brand] = {}

    def resolve_brand(name: str) -> Brand | None:
        key = name.strip()
        if not key:
            return None
        cached = brand_cache.get(key.lower())
        if cached is not None:
            return cached
        # Case-insensitive match so "Samsung", "SAMSUNG", "samsung" all map to
        # the same brand row.
        brand = Brand.objects.filter(name__iexact=key).first()
        if brand is None:
            brand = Brand.objects.create(name=key.upper())
            report.brands_created += 1
            report.warnings.append(f"Created new brand '{brand.name}' (id={brand.id})")
        brand_cache[key.lower()] = brand
        return brand

    for row_index in range(header_index + 1, len(rows)):
        row = rows[row_index]
        if len(row) <= model_col:
            report.skipped_rows += 1
            continue
        model_cell = row[model_col]
        if not model_cell or not str(model_cell).strip():
            report.skipped_rows += 1
            continue

        model_name = str(model_cell).strip()
        seen_model_names.add(model_name)

        # Resolve the brand for this row (if a BRAND column exists and has a
        # value). Auto-creates unknown brands with the next id.
        brand = None
        if brand_col is not None and brand_col < len(row):
            brand_cell = row[brand_col]
            if brand_cell and str(brand_cell).strip():
                brand = resolve_brand(str(brand_cell))

        product, created = Product.objects.get_or_create(model_name=model_name)
        if created:
            report.products_created += 1
        else:
            report.products_updated += 1

        # Update the brand link if we resolved one. We only overwrite when we
        # have a brand from the Excel, so a blank brand cell won't wipe an
        # existing link.
        if brand is not None and product.brand_id != brand.id:
            product.brand = brand
            product.save(update_fields=["brand", "updated_at"])

        # Track which website columns this product actually has a URL for, so
        # in sync mode we can drop sources that were removed from the Excel.
        seen_websites_for_product: set[str] = set()

        for col_index, website_name in website_columns.items():
            if col_index >= len(row):
                continue
            url = _normalize_url(row[col_index])
            if not url:
                continue

            seen_websites_for_product.add(website_name)

            existing = ProductSource.objects.filter(
                product=product, website_name=website_name
            )
            current = existing.filter(url=url).first()
            stale = existing.exclude(url=url)

            if current:
                report.sources_updated += 1
            else:
                stale_first = stale.first()
                if stale_first is not None:
                    stale_first.url = url
                    stale_first.latest_price = None
                    stale_first.availability_status = "unknown"
                    stale_first.crawl_status = "pending"
                    stale_first.error_message = ""
                    stale_first.save(update_fields=[
                        "url", "latest_price", "availability_status",
                        "crawl_status", "error_message", "updated_at",
                    ])
                    stale = stale.exclude(pk=stale_first.pk)
                    report.sources_updated += 1
                else:
                    ProductSource.objects.create(
                        product=product, website_name=website_name, url=url
                    )
                    report.sources_created += 1

            deleted_count = stale.count()
            if deleted_count:
                stale.delete()
                report.sources_deleted += deleted_count
                report.warnings.append(
                    f"Removed {deleted_count} stale URL(s) for "
                    f"{model_name} @ {website_name}"
                )

        # In sync mode, remove sources for this product whose website column
        # was blanked out in the Excel (URL deleted from the sheet).
        if sync:
            orphan_sources = ProductSource.objects.filter(product=product).exclude(
                website_name__in=seen_websites_for_product
            )
            orphan_count = orphan_sources.count()
            if orphan_count:
                orphan_sources.delete()
                report.sources_deleted += orphan_count
                report.warnings.append(
                    f"Removed {orphan_count} source(s) no longer in Excel for {model_name}"
                )

    # SYNC: delete products that are no longer present in the spreadsheet, so
    # the database mirrors the Excel exactly.
    if sync:
        to_delete = Product.objects.exclude(model_name__in=seen_model_names)
        names = list(to_delete.values_list("model_name", flat=True))
        count = to_delete.count()
        if count:
            to_delete.delete()
            report.products_deleted += count
            preview = ", ".join(names[:10]) + ("..." if len(names) > 10 else "")
            report.warnings.append(
                f"Sync: removed {count} product(s) not in Excel: {preview}"
            )

    logger.info("Excel import done: %s", report.as_dict())
    return report
