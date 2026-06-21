"""Excel report builder.

Generates a multi-sheet workbook for the user to download:

    Sheet 1: Summary
        One row per product. Columns: Model, DKP (Digikala product code),
        one column per website (showing latest price), Cheapest site,
        Lowest price, Highest price, Spread, Diff vs cheapest %,
        Last update, Has errors.

    Sheet 2: Sources
        One row per ProductSource (product x website combo). Columns:
        Model, DKP, Website, URL, Latest price, Currency, Availability,
        Crawl status, Last crawled, Error message.

    Sheet 3: Price history
        One row per recorded price observation, useful for charting in Excel.

All sheets use a consistent Arial font, frozen headers, AutoFilter, sensible
column widths and number formats.
"""
from __future__ import annotations

import io
import re
from datetime import datetime

import jdatetime
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from ..models import PriceHistory, Product, ProductSource


# --- Style constants ----------------------------------------------------
HEADER_FONT = Font(name="Arial", bold=True, size=11, color="FFFFFFFF")
HEADER_FILL = PatternFill("solid", start_color="22211C")
HEADER_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)
BODY_FONT = Font(name="Arial", size=11)
PRICE_FMT = "#,##0;[Red](#,##0);-"
CHEAPEST_FILL = PatternFill("solid", start_color="D1FAE5")

# DKP link base — we write a hyperlink formula so the DKP cell is clickable.
DIGIKALA_PRODUCT_BASE = "https://www.digikala.com/product/"

# Regex to extract "dkp-XXXXXXX" from a Digikala URL.
_DKP_RE = re.compile(r"dkp-\d+", re.IGNORECASE)


# --- Helpers ------------------------------------------------------------

def _extract_dkp(url: str) -> str:
    """Return the DKP code from a Digikala URL, e.g. 'dkp-20109389', or ''."""
    if not url:
        return ""
    m = _DKP_RE.search(url)
    return m.group() if m else ""


def _dkp_from_product(product) -> str:
    """Find the DKP code for a product by looking at its Digikala source URL."""
    for src in product.sources.all():
        if "digikala" in src.website_name.lower() or "digikala" in src.url.lower():
            dkp = _extract_dkp(src.url)
            if dkp:
                return dkp
    return ""


def _style_header_row(sheet, row_index: int, last_col: int) -> None:
    for col in range(1, last_col + 1):
        cell = sheet.cell(row=row_index, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
    sheet.row_dimensions[row_index].height = 28


def _autosize_columns(sheet, widths: dict[int, int]) -> None:
    for col_idx, width in widths.items():
        sheet.column_dimensions[get_column_letter(col_idx)].width = width


def _as_jalali_string(value):
    """Convert a Gregorian datetime to a Jalali (Shamsi) string."""
    if value is None:
        return ""
    if isinstance(value, datetime) and value.tzinfo is not None:
        value = value.astimezone(timezone.get_current_timezone()).replace(tzinfo=None)
    if not isinstance(value, datetime):
        return ""
    jd = jdatetime.datetime.fromgregorian(datetime=value)
    return jd.strftime("%Y/%m/%d %H:%M")


# --- Sheet builders -----------------------------------------------------

def _build_summary_sheet(sheet, products, website_names):
    """One row per product, one column per website, DKP as second column."""
    # Column layout: Model | DKP | <website cols> | Cheapest | stats...
    headers = ["Model", "DKP"]
    headers.extend(website_names)
    headers.extend(
        [
            "Cheapest site",
            "Lowest (T)",
            "Highest (T)",
            "Spread (T)",
            "Discount vs highest",
            "Sources tracked",
            "Last update (Jalali)",
            "Has errors",
        ]
    )
    sheet.append(headers)
    _style_header_row(sheet, 1, len(headers))

    dkp_col = 2                            # column B
    site_col_start = 3                     # column C
    website_index = {name: i for i, name in enumerate(website_names)}
    cheapest_col = site_col_start + len(website_names)
    lowest_col = cheapest_col + 1
    highest_col = lowest_col + 1
    spread_col = highest_col + 1
    discount_col = spread_col + 1
    count_col = discount_col + 1
    updated_col = count_col + 1
    errors_col = updated_col + 1

    for product in products:
        dkp = _dkp_from_product(product)

        # Collect each site's latest price, in-stock only.
        prices_per_site: dict[str, float | None] = {name: None for name in website_names}
        site_availability: dict[str, str] = {name: "" for name in website_names}
        latest_update = None
        had_error = False

        for src in product.sources.all():
            name = src.website_name
            if name not in prices_per_site:
                continue
            if src.crawl_status in ("failed", "blocked"):
                had_error = True
            if src.last_crawled_at and (
                latest_update is None or src.last_crawled_at > latest_update
            ):
                latest_update = src.last_crawled_at
            if src.latest_price is None:
                continue
            current = prices_per_site[name]
            existing_av = site_availability[name]
            new_in_stock = src.availability_status != "out_of_stock"
            existing_in_stock = existing_av != "" and existing_av != "out_of_stock"
            if current is None or (new_in_stock and not existing_in_stock) or (
                (new_in_stock == existing_in_stock) and float(src.latest_price) < current
            ):
                prices_per_site[name] = float(src.latest_price)
                site_availability[name] = src.availability_status

        in_stock_prices: list[float] = []
        cheapest_name = ""
        cheapest_price = None
        for name in website_names:
            price = prices_per_site[name]
            if price is not None and site_availability[name] != "out_of_stock":
                in_stock_prices.append(price)
                if cheapest_price is None or price < cheapest_price:
                    cheapest_price = price
                    cheapest_name = name

        lowest = min(in_stock_prices) if in_stock_prices else None
        highest = max(in_stock_prices) if in_stock_prices else None
        spread = (highest - lowest) if (lowest is not None and highest is not None) else None
        discount = (
            (highest - lowest) / highest
            if (lowest is not None and highest and highest > 0 and highest != lowest)
            else None
        )

        row_values = [product.model_name, dkp]
        for name in website_names:
            price = prices_per_site[name]
            if price is None or site_availability[name] == "out_of_stock":
                row_values.append(None)
            else:
                row_values.append(price)
        row_values.extend(
            [
                cheapest_name or "—",
                lowest,
                highest,
                spread,
                discount,
                product.sources.count(),
                _as_jalali_string(latest_update),
                "yes" if had_error else "no",
            ]
        )
        sheet.append(row_values)

        row_index = sheet.max_row

        # Make the DKP cell a clickable hyperlink to the Digikala product page.
        if dkp:
            dkp_cell = sheet.cell(row=row_index, column=dkp_col)
            dkp_cell.hyperlink = f"{DIGIKALA_PRODUCT_BASE}{dkp}/"
            dkp_cell.font = Font(
                name="Arial", size=11, color="0563C1", underline="single"
            )

        # Highlight the cheapest website cell in green.
        if cheapest_name and cheapest_name in website_index:
            col = site_col_start + website_index[cheapest_name]
            sheet.cell(row=row_index, column=col).fill = CHEAPEST_FILL
            sheet.cell(row=row_index, column=col).font = Font(
                name="Arial", size=11, bold=True
            )

        # Apply number formats to the whole row.
        for col_idx in range(1, errors_col + 1):
            cell = sheet.cell(row=row_index, column=col_idx)
            # Don't overwrite the hyperlink font on the DKP cell.
            if col_idx != dkp_col and not (
                cell.font.bold and cell.fill.start_color.rgb == "FFD1FAE5"
            ):
                cell.font = BODY_FONT
            if site_col_start <= col_idx < cheapest_col:
                cell.number_format = PRICE_FMT
            elif col_idx in (lowest_col, highest_col, spread_col):
                cell.number_format = PRICE_FMT
            elif col_idx == discount_col:
                cell.number_format = "0.0%"

    sheet.freeze_panes = "C2"
    if sheet.max_row >= 1:
        sheet.auto_filter.ref = (
            f"A1:{get_column_letter(errors_col)}{sheet.max_row}"
        )

    widths = {1: 22, dkp_col: 14}
    for i in range(len(website_names)):
        widths[site_col_start + i] = 16
    widths[cheapest_col] = 18
    widths[lowest_col] = 14
    widths[highest_col] = 14
    widths[spread_col] = 12
    widths[discount_col] = 14
    widths[count_col] = 10
    widths[updated_col] = 18
    widths[errors_col] = 11
    _autosize_columns(sheet, widths)


def _build_sources_sheet(sheet, sources):
    headers = [
        "Model",
        "DKP",
        "Website",
        "URL",
        "Latest price (T)",
        "Currency",
        "Availability",
        "Crawl status",
        "Last crawled (Jalali)",
        "Consecutive failures",
        "Error message",
    ]
    sheet.append(headers)
    _style_header_row(sheet, 1, len(headers))

    for src in sources:
        dkp = _extract_dkp(src.url) if "digikala" in src.url.lower() else ""
        sheet.append(
            [
                src.product.model_name,
                dkp,
                src.website_name,
                src.url,
                float(src.latest_price) if src.latest_price is not None else None,
                src.currency,
                src.availability_status,
                src.crawl_status,
                _as_jalali_string(src.last_crawled_at),
                src.consecutive_failures,
                src.error_message[:500] if src.error_message else "",
            ]
        )
        row_index = sheet.max_row
        for col_idx in range(1, len(headers) + 1):
            cell = sheet.cell(row=row_index, column=col_idx)
            cell.font = BODY_FONT
            cell.alignment = Alignment(vertical="top", wrap_text=col_idx == 11)
            if col_idx == 5:
                cell.number_format = PRICE_FMT

        # Make the DKP cell in sources sheet a hyperlink too.
        if dkp:
            dkp_cell = sheet.cell(row=row_index, column=2)
            dkp_cell.hyperlink = f"{DIGIKALA_PRODUCT_BASE}{dkp}/"
            dkp_cell.font = Font(
                name="Arial", size=11, color="0563C1", underline="single"
            )

    sheet.freeze_panes = "A2"
    if sheet.max_row >= 1:
        sheet.auto_filter.ref = (
            f"A1:{get_column_letter(len(headers))}{sheet.max_row}"
        )

    _autosize_columns(
        sheet,
        {1: 20, 2: 14, 3: 16, 4: 60, 5: 16, 6: 9,
         7: 14, 8: 12, 9: 18, 10: 10, 11: 42},
    )


def _build_history_sheet(sheet, history_qs):
    headers = ["Crawled at (Jalali)", "Model", "Website", "Price (T)", "Availability"]
    sheet.append(headers)
    _style_header_row(sheet, 1, len(headers))

    for entry in history_qs:
        sheet.append(
            [
                _as_jalali_string(entry.crawled_at),
                entry.product_source.product.model_name,
                entry.product_source.website_name,
                float(entry.price) if entry.price is not None else None,
                entry.availability_status,
            ]
        )
        row_index = sheet.max_row
        for col_idx in range(1, len(headers) + 1):
            cell = sheet.cell(row=row_index, column=col_idx)
            cell.font = BODY_FONT
            if col_idx == 4:
                cell.number_format = PRICE_FMT

    sheet.freeze_panes = "A2"
    if sheet.max_row >= 1:
        sheet.auto_filter.ref = (
            f"A1:{get_column_letter(len(headers))}{sheet.max_row}"
        )

    _autosize_columns(sheet, {1: 18, 2: 20, 3: 16, 4: 16, 5: 14})


# --- Public entry point -------------------------------------------------

def build_report(products_qs=None) -> bytes:
    """Build an .xlsx workbook in memory and return its bytes."""
    products = (
        products_qs if products_qs is not None else Product.objects.all()
    ).prefetch_related("sources")
    products = list(products)

    website_names: list[str] = []
    seen: set[str] = set()
    for product in products:
        for src in product.sources.all():
            if src.website_name not in seen:
                seen.add(src.website_name)
                website_names.append(src.website_name)

    wb = Workbook()

    summary = wb.active
    summary.title = "Summary"
    _build_summary_sheet(summary, products, website_names)

    sources_sheet = wb.create_sheet("Sources")
    all_sources = (
        ProductSource.objects.select_related("product")
        .filter(product__in=products)
        .order_by("product__model_name", "website_name")
    )
    _build_sources_sheet(sources_sheet, all_sources)

    history_sheet = wb.create_sheet("Price history")
    history_qs = (
        PriceHistory.objects.select_related(
            "product_source", "product_source__product"
        )
        .filter(product_source__product__in=products)
        .order_by("-crawled_at")[:5000]
    )
    _build_history_sheet(history_sheet, history_qs)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
